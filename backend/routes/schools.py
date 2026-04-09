"""
Schools CRM routes - including PRO features.
Uses Supabase (PostgreSQL) for database operations.
Supports multi-contact architecture (1 school : N contacts).
"""
import io
import csv
import re
import logging
import uuid
import json as json_lib
from typing import List, Optional
from datetime import datetime, timezone
from pydantic import BaseModel, EmailStr, Field

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

from models.schemas import School, PropagationRequest
from core.security import get_current_user
from database.supabase import get_db
from database.supabase_repositories import (
    SchoolRepositorySupabase,
    InstitutionRepositorySupabase,
    ProgramRepositorySupabase
)

router = APIRouter(prefix="/schools", tags=["Schools"])
logger = logging.getLogger(__name__)

# Max file size: 10MB
MAX_FILE_SIZE = 10 * 1024 * 1024

# Column mapping (various possible names)
COLUMN_MAPPING = {
    'název školy': ['název školy', 'nazev skoly', 'název', 'nazev', 'škola', 'skola', 'school', 'name'],
    'email': ['email', 'e-mail', 'mail', 'e mail'],
    'telefon': ['telefon', 'phone', 'tel', 'tel.', 'telephone'],
    'město': ['město', 'mesto', 'city', 'obec'],
    'poznámka': ['poznámka', 'poznamka', 'note', 'notes', 'poznámky'],
    'kontaktní osoba': ['kontaktní osoba', 'kontaktni osoba', 'kontakt', 'contact', 'contact person', 'jméno kontaktu', 'jmeno kontaktu'],
    'ičo': ['ičo', 'ico', 'ic', 'ič'],
    'typ školy': ['typ školy', 'typ skoly', 'typ', 'type', 'kategorie', 'tag', 'tagy', 'tags'],
}

# Predefined tags
PREDEFINED_TAGS = ['MŠ', 'ZŠ', 'SŠ', 'VOŠ', 'VŠ', 'Gymnázium', 'ZUŠ', 'DDM', 'Jiné']

# Common email domain typos and their corrections
EMAIL_TYPO_CORRECTIONS = {
    'gmial.com': 'gmail.com',
    'gmal.com': 'gmail.com',
    'gmail.cz': 'gmail.com',
    'gamil.com': 'gmail.com',
    'gnail.com': 'gmail.com',
    'hotmal.com': 'hotmail.com',
    'hotmai.com': 'hotmail.com',
    'hotmial.com': 'hotmail.com',
    'sezanm.cz': 'seznam.cz',
    'seznma.cz': 'seznam.cz',
    'sezam.cz': 'seznam.cz',
    'outlook.cz': 'outlook.com',
}


# ============ Pydantic Models ============

class ImportResult(BaseModel):
    success: bool
    total_rows: int
    imported: int
    updated: int
    skipped: int
    errors: int
    duplicates: int
    new_schools: int
    new_contacts: int
    error_details: List[dict]


class SchoolContactCreate(BaseModel):
    email: str
    name: Optional[str] = None
    phone: Optional[str] = None
    is_primary: bool = False
    notes: Optional[str] = None


class SchoolContactUpdate(BaseModel):
    email: Optional[str] = None
    name: Optional[str] = None
    phone: Optional[str] = None
    status: Optional[str] = None
    is_primary: Optional[bool] = None
    notes: Optional[str] = None


class SchoolContactResponse(BaseModel):
    id: str
    school_id: str
    email: str
    name: Optional[str]
    phone: Optional[str]
    status: str
    is_primary: bool
    email_validated: bool
    suggested_correction: Optional[str] = None
    created_at: Optional[str]


class SchoolWithContacts(BaseModel):
    id: str
    name: str
    city: Optional[str]
    address: Optional[str]
    notes: Optional[str]
    tags: List[str]
    source: str
    booking_count: int
    contacts: List[SchoolContactResponse]
    invalid_contacts_count: int = 0
    created_at: Optional[str]


# ============ Helper Functions ============

def normalize_column_name(name: str) -> Optional[str]:
    """Normalize column name to standard format."""
    name_lower = name.lower().strip()
    for standard_name, variations in COLUMN_MAPPING.items():
        if name_lower in variations:
            return standard_name
    return None


def validate_email(email: str) -> tuple[bool, Optional[str]]:
    """
    Validate email format and check for common typos.
    Returns (is_valid, suggested_correction).
    """
    if not email:
        return False, None
    
    email = email.strip().lower()
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    is_valid = bool(re.match(pattern, email))
    
    # Check for common domain typos
    if is_valid:
        domain = email.split('@')[1] if '@' in email else ''
        if domain in EMAIL_TYPO_CORRECTIONS:
            corrected_domain = EMAIL_TYPO_CORRECTIONS[domain]
            corrected_email = email.replace(domain, corrected_domain)
            return True, corrected_email
    
    return is_valid, None


def parse_tags(tag_string: str) -> List[str]:
    """Parse tags from string (comma or semicolon separated)."""
    if not tag_string:
        return []
    tags = []
    for part in re.split(r'[,;]', tag_string):
        tag = part.strip()
        if tag:
            tags.append(tag)
    return tags


def get_school_key(name: str, city: str) -> str:
    """Generate unique key for school (name + city)."""
    return f"{name.lower().strip()}|{(city or '').lower().strip()}"


# ============ File Parsing ============

def parse_excel_file(file_content: bytes, filename: str) -> tuple[List[dict], List[dict]]:
    """Parse Excel file and return rows with errors."""
    import openpyxl
    
    rows = []
    errors = []
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                normalized = normalize_column_name(str(cell.value))
                headers.append(normalized if normalized else str(cell.value).lower())
            else:
                headers.append(None)
        
        # Check required columns
        if 'název školy' not in headers:
            errors.append({"row": 0, "error": "Chybí povinný sloupec 'Název školy'"})
            return rows, errors
        if 'email' not in headers:
            errors.append({"row": 0, "error": "Chybí povinný sloupec 'Email'"})
            return rows, errors
        if 'město' not in headers:
            errors.append({"row": 0, "error": "Chybí povinný sloupec 'Město' (nutné pro deduplikaci)"})
            return rows, errors
        
        # Parse data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            
            row_data = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    row_data[headers[col_idx]] = str(value).strip() if value and str(value).strip() != 'None' else ''
            
            # Validate required fields
            name = row_data.get('název školy', '').strip()
            city = row_data.get('město', '').strip()
            email = row_data.get('email', '').strip()
            
            if not name:
                errors.append({"row": row_idx, "error": f"Řádek {row_idx}: Chybí název školy"})
                continue
            
            if not city:
                errors.append({"row": row_idx, "error": f"Řádek {row_idx}: Chybí město"})
                continue
            
            if not email:
                errors.append({"row": row_idx, "error": f"Řádek {row_idx}: Chybí email"})
                continue
            
            is_valid, suggested = validate_email(email)
            if not is_valid:
                errors.append({"row": row_idx, "error": f"Řádek {row_idx}: Neplatný formát emailu '{email}'"})
                continue
            
            rows.append({
                'school_name': name,
                'city': city,
                'email': email.lower(),
                'email_suggestion': suggested,
                'phone': row_data.get('telefon', ''),
                'contact_name': row_data.get('kontaktní osoba', ''),
                'notes': row_data.get('poznámka', ''),
                'tags': parse_tags(row_data.get('typ školy', '')),
            })
        
        wb.close()
        
    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        errors.append({"row": 0, "error": f"Chyba při čtení souboru: {str(e)}"})
    
    return rows, errors


def parse_csv_file(file_content: bytes, filename: str) -> tuple[List[dict], List[dict]]:
    """Parse CSV file and return rows with errors."""
    rows = []
    errors = []
    
    try:
        content = file_content.decode('utf-8-sig')
        delimiter = ';' if ';' in content.split('\n')[0] else ','
        
        reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
        
        # Normalize headers
        fieldnames = {}
        for field in reader.fieldnames or []:
            normalized = normalize_column_name(field)
            if normalized:
                fieldnames[field] = normalized
            else:
                fieldnames[field] = field.lower()
        
        normalized_fields = list(fieldnames.values())
        if 'název školy' not in normalized_fields:
            errors.append({"row": 0, "error": "Chybí povinný sloupec 'Název školy'"})
            return rows, errors
        if 'email' not in normalized_fields:
            errors.append({"row": 0, "error": "Chybí povinný sloupec 'Email'"})
            return rows, errors
        if 'město' not in normalized_fields:
            errors.append({"row": 0, "error": "Chybí povinný sloupec 'Město' (nutné pro deduplikaci)"})
            return rows, errors
        
        for row_idx, row in enumerate(reader, start=2):
            normalized_row = {}
            for key, value in row.items():
                norm_key = fieldnames.get(key, key.lower())
                normalized_row[norm_key] = value.strip() if value else ''
            
            name = normalized_row.get('název školy', '').strip()
            city = normalized_row.get('město', '').strip()
            email = normalized_row.get('email', '').strip()
            
            if not name and not email:
                continue
            
            if not name:
                errors.append({"row": row_idx, "error": f"Řádek {row_idx}: Chybí název školy"})
                continue
            
            if not city:
                errors.append({"row": row_idx, "error": f"Řádek {row_idx}: Chybí město"})
                continue
            
            if not email:
                errors.append({"row": row_idx, "error": f"Řádek {row_idx}: Chybí email"})
                continue
            
            is_valid, suggested = validate_email(email)
            if not is_valid:
                errors.append({"row": row_idx, "error": f"Řádek {row_idx}: Neplatný formát emailu '{email}'"})
                continue
            
            rows.append({
                'school_name': name,
                'city': city,
                'email': email.lower(),
                'email_suggestion': suggested,
                'phone': normalized_row.get('telefon', ''),
                'contact_name': normalized_row.get('kontaktní osoba', ''),
                'notes': normalized_row.get('poznámka', ''),
                'tags': parse_tags(normalized_row.get('typ školy', '')),
            })
    
    except UnicodeDecodeError:
        try:
            content = file_content.decode('latin-1')
            return parse_csv_file(content.encode('utf-8'), filename)
        except Exception as e:
            errors.append({"row": 0, "error": f"Chyba kódování souboru: {str(e)}"})
    except Exception as e:
        logger.error(f"Error parsing CSV file: {e}")
        errors.append({"row": 0, "error": f"Chyba při čtení souboru: {str(e)}"})
    
    return rows, errors


# ============ Database Setup ============

@router.post("/setup-contacts-table")
async def setup_contacts_table(
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Create school_contacts table if not exists."""
    if current_user.get("role") not in ["admin", "spravce"]:
        raise HTTPException(status_code=403, detail="Přístup odepřen")
    
    try:
        # Add missing columns to schools table
        try:
            await db.execute(text("""
                ALTER TABLE schools 
                ADD COLUMN IF NOT EXISTS tags JSONB DEFAULT '[]'::jsonb,
                ADD COLUMN IF NOT EXISTS source TEXT DEFAULT 'organic'
            """))
        except Exception:
            pass  # Columns may already exist
        
        # Create school_contacts table
        await db.execute(text("""
            CREATE TABLE IF NOT EXISTS school_contacts (
                id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                school_id UUID NOT NULL REFERENCES schools(id) ON DELETE CASCADE,
                institution_id UUID NOT NULL REFERENCES institutions(id) ON DELETE CASCADE,
                email TEXT NOT NULL,
                name TEXT,
                phone TEXT,
                status TEXT NOT NULL DEFAULT 'active',
                email_validated BOOLEAN DEFAULT FALSE,
                email_validation_error TEXT,
                last_email_sent_at TIMESTAMPTZ,
                last_email_bounced BOOLEAN DEFAULT FALSE,
                is_primary BOOLEAN DEFAULT FALSE,
                notes TEXT,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
        """))
        
        # Create indexes one by one
        index_queries = [
            "CREATE INDEX IF NOT EXISTS idx_school_contacts_school ON school_contacts(school_id)",
            "CREATE INDEX IF NOT EXISTS idx_school_contacts_email ON school_contacts(email)",
            "CREATE INDEX IF NOT EXISTS idx_school_contacts_institution ON school_contacts(institution_id)",
            "CREATE INDEX IF NOT EXISTS idx_school_contacts_status ON school_contacts(status)",
            "CREATE INDEX IF NOT EXISTS idx_schools_name_city ON schools(institution_id, name, city)"
        ]
        
        for query in index_queries:
            try:
                await db.execute(text(query))
            except Exception:
                pass  # Index may already exist
        
        await db.commit()
        return {"message": "Tabulka school_contacts vytvořena", "success": True}
        
    except Exception as e:
        logger.error(f"Error setting up contacts table: {e}")
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/migrate-contacts")
async def migrate_existing_contacts(
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Migrate existing school emails to school_contacts table."""
    if current_user.get("role") not in ["admin", "spravce"]:
        raise HTTPException(status_code=403, detail="Přístup odepřen")
    
    try:
        # Get schools with email but no contacts
        result = await db.execute(text("""
            SELECT s.id, s.institution_id, s.email, s.contact_person, s.phone
            FROM schools s
            WHERE s.email IS NOT NULL 
            AND s.email != ''
            AND s.institution_id = :inst_id
            AND NOT EXISTS (
                SELECT 1 FROM school_contacts sc WHERE sc.school_id = s.id
            )
        """), {"inst_id": current_user["institution_id"]})
        
        schools_to_migrate = result.fetchall()
        migrated = 0
        
        for school in schools_to_migrate:
            await db.execute(text("""
                INSERT INTO school_contacts (id, school_id, institution_id, email, name, phone, is_primary, status)
                VALUES (:id, :school_id, :inst_id, :email, :name, :phone, TRUE, 'active')
            """), {
                "id": str(uuid.uuid4()),
                "school_id": str(school[0]),
                "inst_id": str(school[1]),
                "email": school[2],
                "name": school[3] or '',
                "phone": school[4] or ''
            })
            migrated += 1
        
        await db.commit()
        return {"message": f"Migrace dokončena", "migrated": migrated}
        
    except Exception as e:
        logger.error(f"Error migrating contacts: {e}")
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ============ Static Routes (MUST be before /{school_id} route) ============

@router.get("/tags")
async def get_all_tags(
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get all unique tags used by schools in institution."""
    result = await db.execute(text("""
        SELECT DISTINCT jsonb_array_elements_text(tags) as tag
        FROM schools
        WHERE institution_id = :inst_id AND tags IS NOT NULL AND jsonb_array_length(tags) > 0
        ORDER BY tag
    """), {"inst_id": current_user["institution_id"]})
    
    tags = [row[0] for row in result.fetchall()]
    return {"tags": tags}


@router.get("/import-template")
async def download_import_template():
    """Download sample Excel template for school import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Školy"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B3E50", end_color="2B3E50", fill_type="solid")
    required_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers - NEW FORMAT with Město as required (IČO removed)
    headers = ["Název školy", "Město", "Email", "Jméno kontaktu", "Telefon", "Typ školy", "Poznámka"]
    required_cols = [0, 1, 2]  # Name, City, Email are required
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
    
    # Sample data - demonstrating multiple contacts per school
    sample_data = [
        ["ZŠ Příkladná", "Praha", "info@zsprikladna.cz", "Sekretariát", "+420 123 456 789", "ZŠ", "Hlavní kontakt"],
        ["ZŠ Příkladná", "Praha", "ucitel.novak@zsprikladna.cz", "Jan Novák", "721 000 111", "", "Pedagog - přírodovědné programy"],
        ["ZŠ Příkladná", "Praha", "ucitelka.svoboda@zsprikladna.cz", "Marie Svobodová", "", "", "Pedagog - výtvarné programy"],
        ["MŠ Sluníčko", "Brno", "skolka@slunicko.cz", "Ředitelka", "722 333 444", "MŠ", ""],
        ["Gymnázium ABC", "Ostrava", "sekretariat@gymnabc.cz", "", "", "Gymnázium; SŠ", "Zájem o přírodovědné programy"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx - 1 in required_cols:
                cell.fill = required_fill
    
    # Instructions sheet
    ws_info = wb.create_sheet("Nápověda")
    instructions = [
        ["INSTRUKCE PRO IMPORT ŠKOL A KONTAKTŮ"],
        [""],
        ["🔴 Povinné sloupce (žluté):"],
        ["  - Název školy: Celý název školy"],
        ["  - Město: Město/obec (nutné pro deduplikaci)"],
        ["  - Email: Platná emailová adresa kontaktu"],
        [""],
        ["📧 MULTI-KONTAKT SYSTÉM:"],
        ["  - Jedna škola může mít více kontaktů (pedagogů)"],
        ["  - Škola je identifikována kombinací Název + Město"],
        ["  - Pokud škola existuje, přidá se pouze nový kontakt"],
        ["  - Duplikátní emaily jsou přeskočeny"],
        [""],
        ["Příklad:"],
        ["  ZŠ Příkladná | Praha | info@skola.cz (hlavní kontakt)"],
        ["  ZŠ Příkladná | Praha | ucitel@skola.cz (další pedagog)"],
        ["  → Vytvoří 1 školu s 2 kontakty"],
        [""],
        ["Volitelné sloupce:"],
        ["  - Jméno kontaktu: Jméno kontaktní osoby"],
        ["  - Telefon: Telefonní číslo"],
        ["  - Typ školy: Kategorie (MŠ, ZŠ, SŠ, atd.) - více oddělte čárkou"],
        ["  - Poznámka: Libovolná poznámka"],
        [""],
        ["Dostupné typy škol (tagy):"],
        ["  MŠ, ZŠ, SŠ, VOŠ, VŠ, Gymnázium, ZUŠ, DDM, Jiné"],
        [""],
        ["Důležité:"],
        ["  - První řádek musí obsahovat názvy sloupců"],
        ["  - Maximální velikost souboru: 10 MB"],
        ["  - Podporované formáty: .xlsx, .xls, .csv"],
    ]
    
    for row_idx, row in enumerate(instructions, 1):
        cell = ws_info.cell(row=row_idx, column=1, value=row[0] if row else "")
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        ws_info.column_dimensions['A'].width = 65
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vzorovy_import_skol_kontaktu.xlsx"}
    )


@router.get("/export-csv")
async def export_schools_csv(
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Export schools with contacts to CSV - PRO feature only."""
    institution_repo = InstitutionRepositorySupabase(db)
    
    institution = await institution_repo.find_by_id(current_user["institution_id"])
    if not institution or institution.get("plan") not in ["standard", "premium"]:
        raise HTTPException(status_code=403, detail="Tato funkce je dostupná pouze v PRO verzi")
    
    # Get schools with contacts
    result = await db.execute(text("""
        SELECT s.name, s.city, sc.email, sc.name as contact_name, sc.phone, sc.status, s.created_at
        FROM schools s
        LEFT JOIN school_contacts sc ON sc.school_id = s.id
        WHERE s.institution_id = :inst_id AND s.deleted_at IS NULL
        ORDER BY s.name, sc.is_primary DESC
    """), {"inst_id": current_user["institution_id"]})
    
    output = io.StringIO()
    output.write("Název školy;Město;Email;Kontaktní osoba;Telefon;Status;Datum registrace\n")
    
    for row in result.fetchall():
        created = row[6].strftime("%d.%m.%Y") if row[6] else ""
        output.write(f"{row[0]};{row[1] or ''};{row[2] or ''};{row[3] or ''};{row[4] or ''};{row[5] or ''};{created}\n")
    
    csv_content = output.getvalue()
    output.close()
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=skoly_kontakty_export.csv"}
    )


@router.get("/campaign-contacts")
async def get_campaign_contacts(
    tag: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get all active, valid email contacts for campaigns."""
    query = """
        SELECT DISTINCT sc.email, sc.name, s.name as school_name
        FROM school_contacts sc
        JOIN schools s ON s.id = sc.school_id
        WHERE sc.institution_id = :inst_id 
        AND sc.status = 'active'
        AND s.deleted_at IS NULL
    """
    params = {"inst_id": current_user["institution_id"]}
    
    if tag:
        query += " AND s.tags @> CAST(:tag AS jsonb)"
        params["tag"] = json_lib.dumps([tag])
    
    query += " ORDER BY s.name, sc.name"
    
    result = await db.execute(text(query), params)
    
    contacts = []
    for row in result.fetchall():
        contacts.append({
            "email": row[0],
            "name": row[1],
            "school_name": row[2]
        })
    
    return {"contacts": contacts, "total": len(contacts)}


# ============ Schools API ============

@router.get("")
async def get_schools(
    source: Optional[str] = None,
    tag: Optional[str] = None,
    has_invalid: Optional[bool] = None,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get all schools with their contacts."""
    inst_id = current_user["institution_id"]
    
    # Build query
    query = """
        SELECT 
            s.id, s.name, s.city, s.address, s.notes, s.tags, s.source, 
            s.booking_count, s.created_at, s.email as legacy_email, 
            s.contact_person as legacy_contact, s.phone as legacy_phone
        FROM schools s
        WHERE s.institution_id = :inst_id AND s.deleted_at IS NULL
    """
    params = {"inst_id": inst_id}
    
    if source and source != 'all':
        query += " AND s.source = :source"
        params["source"] = source
    
    if tag and tag != 'all':
        # Use CAST instead of :: to avoid parameter parsing issues
        query += " AND s.tags @> CAST(:tag AS jsonb)"
        params["tag"] = json_lib.dumps([tag])
    
    query += " ORDER BY s.name ASC"
    
    result = await db.execute(text(query), params)
    schools_raw = result.fetchall()
    
    # Get all contacts for these schools
    school_ids = [str(s[0]) for s in schools_raw]
    contacts_map = {}
    
    if school_ids:
        # Parametrized IN clause — safe against SQL injection
        id_params = {f"sid_{i}": sid for i, sid in enumerate(school_ids)}
        id_placeholders = ", ".join(f":sid_{i}" for i in range(len(school_ids)))
        contacts_result = await db.execute(text(f"""
            SELECT id, school_id, email, name, phone, status, is_primary, email_validated, created_at
            FROM school_contacts
            WHERE school_id::text IN ({id_placeholders})
            ORDER BY is_primary DESC, created_at ASC
        """), id_params)
        
        for c in contacts_result.fetchall():
            school_id = str(c[1])
            if school_id not in contacts_map:
                contacts_map[school_id] = []
            
            # Check for email typo suggestion
            _, suggested = validate_email(c[2])
            
            contacts_map[school_id].append({
                "id": str(c[0]),
                "school_id": school_id,
                "email": c[2],
                "name": c[3],
                "phone": c[4],
                "status": c[5],
                "is_primary": c[6],
                "email_validated": c[7] or False,
                "suggested_correction": suggested,
                "created_at": c[8].isoformat() if c[8] else None
            })
    
    # Build response
    schools = []
    for s in schools_raw:
        school_id = str(s[0])
        contacts = contacts_map.get(school_id, [])
        
        # If no contacts, create one from legacy fields
        if not contacts and s[9]:  # legacy_email exists (index 9)
            contacts = [{
                "id": None,
                "school_id": school_id,
                "email": s[9],
                "name": s[10],  # legacy_contact
                "phone": s[11],  # legacy_phone
                "status": "active",
                "is_primary": True,
                "email_validated": False,
                "suggested_correction": None,
                "created_at": None
            }]
        
        tags = s[5] if s[5] else []
        if isinstance(tags, str):
            try:
                tags = json_lib.loads(tags)
            except:
                tags = []
        
        invalid_count = len([c for c in contacts if c.get("status") == "invalid"])
        
        # Filter by has_invalid if specified
        if has_invalid is not None:
            if has_invalid and invalid_count == 0:
                continue
            if not has_invalid and invalid_count > 0:
                continue
        
        schools.append({
            "id": school_id,
            "name": s[1],
            "city": s[2],
            "address": s[3],
            "notes": s[4],
            "tags": tags,
            "source": s[6] or "organic",
            "booking_count": s[7] or 0,
            "contacts": contacts,
            "invalid_contacts_count": invalid_count,
            "created_at": s[8].isoformat() if s[8] else None,
            # Legacy fields for backward compatibility
            "email": contacts[0]["email"] if contacts else s[9],
            "contact_person": contacts[0]["name"] if contacts else s[10],
            "phone": contacts[0]["phone"] if contacts else s[11],
        })
    
    return schools


@router.get("/{school_id}")
async def get_school(
    school_id: str,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get single school with all contacts."""
    result = await db.execute(text("""
        SELECT id, name, city, address, notes, tags, source, booking_count, created_at
        FROM schools
        WHERE id = :id AND institution_id = :inst_id AND deleted_at IS NULL
    """), {"id": school_id, "inst_id": current_user["institution_id"]})
    
    school = result.fetchone()
    if not school:
        raise HTTPException(status_code=404, detail="Škola nenalezena")
    
    # Get contacts
    contacts_result = await db.execute(text("""
        SELECT id, school_id, email, name, phone, status, is_primary, email_validated, created_at
        FROM school_contacts
        WHERE school_id = :school_id
        ORDER BY is_primary DESC, created_at ASC
    """), {"school_id": school_id})
    
    contacts = []
    for c in contacts_result.fetchall():
        _, suggested = validate_email(c[2])
        contacts.append({
            "id": str(c[0]),
            "school_id": str(c[1]),
            "email": c[2],
            "name": c[3],
            "phone": c[4],
            "status": c[5],
            "is_primary": c[6],
            "email_validated": c[7] or False,
            "suggested_correction": suggested,
            "created_at": c[8].isoformat() if c[8] else None
        })
    
    tags = school[5] if school[5] else []
    if isinstance(tags, str):
        try:
            tags = json_lib.loads(tags)
        except:
            tags = []
    
    return {
        "id": str(school[0]),
        "name": school[1],
        "city": school[2],
        "address": school[3],
        "notes": school[4],
        "tags": tags,
        "source": school[6] or "organic",
        "booking_count": school[7] or 0,
        "contacts": contacts,
        "invalid_contacts_count": len([c for c in contacts if c["status"] == "invalid"]),
        "created_at": school[8].isoformat() if school[8] else None
    }


# ============ Contacts API ============

@router.post("/{school_id}/contacts")
async def add_contact(
    school_id: str,
    contact: SchoolContactCreate,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Add a new contact to a school."""
    # Verify school belongs to institution
    result = await db.execute(text("""
        SELECT id FROM schools 
        WHERE id = :school_id AND institution_id = :inst_id
    """), {"school_id": school_id, "inst_id": current_user["institution_id"]})
    
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Škola nenalezena")
    
    # Validate email
    is_valid, suggested = validate_email(contact.email)
    if not is_valid:
        raise HTTPException(status_code=400, detail="Neplatný formát emailu")
    
    # Check for duplicate email in institution
    existing = await db.execute(text("""
        SELECT id FROM school_contacts 
        WHERE email = :email AND institution_id = :inst_id
    """), {"email": contact.email.lower(), "inst_id": current_user["institution_id"]})
    
    if existing.fetchone():
        raise HTTPException(status_code=400, detail="Email již existuje v jiné škole")
    
    # If this is primary, unset other primary contacts
    if contact.is_primary:
        await db.execute(text("""
            UPDATE school_contacts SET is_primary = FALSE 
            WHERE school_id = :school_id
        """), {"school_id": school_id})
    
    # Insert contact
    contact_id = str(uuid.uuid4())
    await db.execute(text("""
        INSERT INTO school_contacts (id, school_id, institution_id, email, name, phone, is_primary, notes, status)
        VALUES (:id, :school_id, :inst_id, :email, :name, :phone, :is_primary, :notes, 'active')
    """), {
        "id": contact_id,
        "school_id": school_id,
        "inst_id": current_user["institution_id"],
        "email": contact.email.lower(),
        "name": contact.name or '',
        "phone": contact.phone or '',
        "is_primary": contact.is_primary,
        "notes": contact.notes or ''
    })
    
    await db.commit()
    
    return {
        "id": contact_id,
        "message": "Kontakt přidán",
        "suggested_correction": suggested
    }


@router.put("/{school_id}/contacts/{contact_id}")
async def update_contact(
    school_id: str,
    contact_id: str,
    contact: SchoolContactUpdate,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Update a contact."""
    # Verify contact belongs to school and institution
    result = await db.execute(text("""
        SELECT sc.id FROM school_contacts sc
        JOIN schools s ON s.id = sc.school_id
        WHERE sc.id = :contact_id AND sc.school_id = :school_id AND s.institution_id = :inst_id
    """), {"contact_id": contact_id, "school_id": school_id, "inst_id": current_user["institution_id"]})
    
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Kontakt nenalezen")
    
    updates = []
    params = {"contact_id": contact_id, "updated": datetime.now(timezone.utc)}
    
    if contact.email is not None:
        is_valid, _ = validate_email(contact.email)
        if not is_valid:
            raise HTTPException(status_code=400, detail="Neplatný formát emailu")
        updates.append("email = :email")
        params["email"] = contact.email.lower()
    
    if contact.name is not None:
        updates.append("name = :name")
        params["name"] = contact.name
    
    if contact.phone is not None:
        updates.append("phone = :phone")
        params["phone"] = contact.phone
    
    if contact.status is not None:
        if contact.status not in ['active', 'invalid', 'pending_verification']:
            raise HTTPException(status_code=400, detail="Neplatný status")
        updates.append("status = :status")
        params["status"] = contact.status
    
    if contact.is_primary is not None:
        if contact.is_primary:
            await db.execute(text("""
                UPDATE school_contacts SET is_primary = FALSE 
                WHERE school_id = :school_id
            """), {"school_id": school_id})
        updates.append("is_primary = :is_primary")
        params["is_primary"] = contact.is_primary
    
    if contact.notes is not None:
        updates.append("notes = :notes")
        params["notes"] = contact.notes
    
    updates.append("updated_at = :updated")
    
    if updates:
        await db.execute(text(f"""
            UPDATE school_contacts SET {', '.join(updates)}
            WHERE id = :contact_id
        """), params)
        await db.commit()
    
    return {"message": "Kontakt aktualizován"}


@router.delete("/{school_id}/contacts/{contact_id}")
async def delete_contact(
    school_id: str,
    contact_id: str,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Delete a contact."""
    result = await db.execute(text("""
        SELECT sc.id FROM school_contacts sc
        JOIN schools s ON s.id = sc.school_id
        WHERE sc.id = :contact_id AND sc.school_id = :school_id AND s.institution_id = :inst_id
    """), {"contact_id": contact_id, "school_id": school_id, "inst_id": current_user["institution_id"]})
    
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Kontakt nenalezen")
    
    await db.execute(text("DELETE FROM school_contacts WHERE id = :id"), {"id": contact_id})
    await db.commit()
    
    return {"message": "Kontakt odstraněn"}


@router.post("/{school_id}/contacts/{contact_id}/fix-email")
async def fix_contact_email(
    school_id: str,
    contact_id: str,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Auto-fix email typo if suggestion available."""
    result = await db.execute(text("""
        SELECT sc.email FROM school_contacts sc
        JOIN schools s ON s.id = sc.school_id
        WHERE sc.id = :contact_id AND sc.school_id = :school_id AND s.institution_id = :inst_id
    """), {"contact_id": contact_id, "school_id": school_id, "inst_id": current_user["institution_id"]})
    
    row = result.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Kontakt nenalezen")
    
    _, suggested = validate_email(row[0])
    if not suggested:
        raise HTTPException(status_code=400, detail="Žádná oprava není k dispozici")
    
    # Check if corrected email already exists
    existing = await db.execute(text("""
        SELECT id FROM school_contacts 
        WHERE email = :email AND institution_id = :inst_id AND id != :contact_id
    """), {"email": suggested, "inst_id": current_user["institution_id"], "contact_id": contact_id})
    
    if existing.fetchone():
        raise HTTPException(status_code=400, detail="Opravený email již existuje v systému")
    
    await db.execute(text("""
        UPDATE school_contacts 
        SET email = :email, status = 'active', updated_at = :updated
        WHERE id = :contact_id
    """), {"email": suggested, "contact_id": contact_id, "updated": datetime.now(timezone.utc)})
    
    await db.commit()
    
    return {"message": "Email opraven", "new_email": suggested}


# ============ Tags Update API ============

@router.put("/{school_id}/tags")
async def update_school_tags(
    school_id: str,
    tags: List[str],
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Update tags for a school."""
    result = await db.execute(text("""
        SELECT id FROM schools 
        WHERE id = :school_id AND institution_id = :inst_id
    """), {"school_id": school_id, "inst_id": current_user["institution_id"]})
    
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Škola nenalezena")
    
    await db.execute(text("""
        UPDATE schools SET tags = :tags, updated_at = :updated
        WHERE id = :school_id
    """), {
        "tags": json_lib.dumps(tags),
        "school_id": school_id,
        "updated": datetime.now(timezone.utc)
    })
    await db.commit()
    
    return {"message": "Tagy aktualizovány", "tags": tags}


@router.post("/import", response_model=ImportResult)
async def import_schools(
    file: UploadFile = File(...),
    update_existing: bool = False,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Import schools and contacts from Excel or CSV file.
    
    Deduplication logic:
    - School: unique by (name + city)
    - Contact: unique by email across institution
    """
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="Soubor je příliš velký (max 10 MB)")
    
    filename = file.filename.lower()
    if not any(filename.endswith(ext) for ext in ['.xlsx', '.xls', '.csv']):
        raise HTTPException(status_code=400, detail="Nepodporovaný formát souboru")
    
    # Parse file
    if filename.endswith('.csv'):
        rows, errors = parse_csv_file(content, filename)
    else:
        rows, errors = parse_excel_file(content, filename)
    
    if not rows and errors:
        return ImportResult(
            success=False, total_rows=0, imported=0, updated=0, skipped=0,
            errors=len(errors), duplicates=0, new_schools=0, new_contacts=0,
            error_details=errors[:100]
        )
    
    inst_id = current_user["institution_id"]
    
    # Get existing schools (by name + city)
    existing_schools_result = await db.execute(text("""
        SELECT id, name, city FROM schools 
        WHERE institution_id = :inst_id AND deleted_at IS NULL
    """), {"inst_id": inst_id})
    
    existing_schools = {}
    for row in existing_schools_result.fetchall():
        key = get_school_key(row[1], row[2])
        existing_schools[key] = str(row[0])
    
    # Get existing emails
    existing_emails_result = await db.execute(text("""
        SELECT email FROM school_contacts WHERE institution_id = :inst_id
    """), {"inst_id": inst_id})
    existing_emails = {row[0].lower() for row in existing_emails_result.fetchall()}
    
    # Also check legacy emails
    legacy_emails_result = await db.execute(text("""
        SELECT email FROM schools WHERE institution_id = :inst_id AND email IS NOT NULL
    """), {"inst_id": inst_id})
    existing_emails.update({row[0].lower() for row in legacy_emails_result.fetchall() if row[0]})
    
    # Process rows
    new_schools = 0
    new_contacts = 0
    updated = 0
    duplicates = 0
    
    # Group rows by school
    school_data = {}
    for row in rows:
        school_key = get_school_key(row['school_name'], row['city'])
        
        if school_key not in school_data:
            school_data[school_key] = {
                'name': row['school_name'],
                'city': row['city'],
                'notes': row.get('notes', ''),
                'tags': row.get('tags', []),
                'contacts': []
            }
        
        # Add contact if email not duplicate
        email = row['email'].lower()
        if email not in existing_emails:
            school_data[school_key]['contacts'].append({
                'email': email,
                'name': row.get('contact_name', ''),
                'phone': row.get('phone', ''),
                'suggestion': row.get('email_suggestion')
            })
            existing_emails.add(email)  # Track within import
        else:
            duplicates += 1
    
    # Insert/update schools and contacts
    for school_key, data in school_data.items():
        school_id = existing_schools.get(school_key)
        
        if not school_id:
            # Create new school
            school_id = str(uuid.uuid4())
            await db.execute(text("""
                INSERT INTO schools (id, institution_id, name, city, notes, tags, source, booking_count, created_at, updated_at)
                VALUES (:id, :inst_id, :name, :city, :notes, :tags, 'import', 0, :created, :updated)
            """), {
                "id": school_id,
                "inst_id": inst_id,
                "name": data['name'],
                "city": data['city'],
                "notes": data['notes'],
                "tags": json_lib.dumps(data['tags']),
                "created": datetime.now(timezone.utc),
                "updated": datetime.now(timezone.utc)
            })
            existing_schools[school_key] = school_id
            new_schools += 1
        elif update_existing:
            # Update school tags/notes
            await db.execute(text("""
                UPDATE schools SET 
                    tags = COALESCE(tags, '[]'::jsonb) || :new_tags::jsonb,
                    updated_at = :updated
                WHERE id = :id
            """), {
                "id": school_id,
                "new_tags": json_lib.dumps(data['tags']),
                "updated": datetime.now(timezone.utc)
            })
            updated += 1
        
        # Add contacts
        is_first = True
        for contact in data['contacts']:
            contact_id = str(uuid.uuid4())
            await db.execute(text("""
                INSERT INTO school_contacts (id, school_id, institution_id, email, name, phone, is_primary, status)
                VALUES (:id, :school_id, :inst_id, :email, :name, :phone, :is_primary, 'active')
            """), {
                "id": contact_id,
                "school_id": school_id,
                "inst_id": inst_id,
                "email": contact['email'],
                "name": contact['name'],
                "phone": contact['phone'],
                "is_primary": is_first  # First contact is primary
            })
            new_contacts += 1
            is_first = False
    
    await db.commit()
    
    logger.info(f"Import completed: {new_schools} schools, {new_contacts} contacts, {duplicates} duplicates")
    
    return ImportResult(
        success=new_schools > 0 or new_contacts > 0,
        total_rows=len(rows),
        imported=new_schools + new_contacts,
        updated=updated,
        skipped=0,
        errors=len(errors),
        duplicates=duplicates,
        new_schools=new_schools,
        new_contacts=new_contacts,
        error_details=errors[:100]
    )


@router.post("/send-propagation")
async def send_propagation(
    request: PropagationRequest,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Send program promotion to school contacts - PRO feature only."""
    institution_repo = InstitutionRepositorySupabase(db)
    program_repo = ProgramRepositorySupabase(db)
    
    institution = await institution_repo.find_by_id(current_user["institution_id"])
    if not institution or institution.get("plan") not in ["standard", "premium"]:
        raise HTTPException(status_code=403, detail="Tato funkce je dostupná pouze v PRO verzi")
    
    program = await program_repo.find_by_id(request.program_id, current_user["institution_id"])
    if not program:
        raise HTTPException(status_code=404, detail="Program nenalezen")
    
    # Get active contacts for selected schools (parametrized)
    id_params = {f"sid_{i}": sid for i, sid in enumerate(request.school_ids)}
    id_params["inst_id"] = current_user["institution_id"]
    id_placeholders = ", ".join(f":sid_{i}" for i in range(len(request.school_ids)))
    result = await db.execute(text(f"""
        SELECT DISTINCT sc.email, sc.name, s.name as school_name
        FROM school_contacts sc
        JOIN schools s ON s.id = sc.school_id
        WHERE s.id::text IN ({id_placeholders})
        AND sc.status = 'active'
        AND s.institution_id = :inst_id
    """), id_params)
    
    contacts = result.fetchall()
    
    if not contacts:
        raise HTTPException(status_code=400, detail="Žádné aktivní kontakty k odeslání")
    
    # Format email
    pro_settings = institution.get("pro_settings", {}) or {}
    subject = pro_settings.get("email_subject_template", "Nový program: {program_name}")
    body = pro_settings.get("email_body_template", 
        "Dobrý den,\n\nrádi bychom Vás informovali o novém programu {program_name}.\n\n"
        "{program_description}\n\nRezervovat můžete zde: {reservation_url}\n\n"
        "S pozdravem,\n{institution_name}"
    )
    
    reservation_url = f"https://www.budezivo.cz/booking/{current_user['institution_id']}?program={request.program_id}"
    
    subject = subject.replace("{program_name}", program.get("name_cs", ""))
    body = body.replace("{program_name}", program.get("name_cs", ""))
    body = body.replace("{program_description}", program.get("description_cs", ""))
    body = body.replace("{reservation_url}", reservation_url)
    body = body.replace("{institution_name}", institution.get("name", ""))
    
    # Send emails (mock for now)
    sent_count = 0
    for contact in contacts:
        logger.info(f"[PROPAGATION] Sending to {contact[0]}: {subject}")
        sent_count += 1
    
    return {
        "message": f"Propagace odeslána {sent_count} kontaktům",
        "sent_count": sent_count,
        "contacts": [c[0] for c in contacts]
    }
