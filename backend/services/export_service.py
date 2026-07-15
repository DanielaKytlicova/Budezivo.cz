"""
Export service for event applications.
Generates XLSX, CSV, and PDF exports.
"""
import io
import csv
import uuid
import logging
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import registerFontFamily

logger = logging.getLogger(__name__)

# ============ Font registration (once, at module load) ============
# Bundled fonts in /app/backend/fonts/ guarantee Czech diacritics work in all PDFs.
_FONT_BASE = 'Helvetica'
_FONT_BOLD = 'Helvetica-Bold'

def _register_pdf_fonts():
    """Register DejaVuSans as the default PDF font family. Supports full Czech diacritics."""
    global _FONT_BASE, _FONT_BOLD
    import os
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    candidates = [
        os.path.join(base_dir, 'fonts'),
        '/usr/share/fonts/truetype/dejavu',
    ]
    for d in candidates:
        regular = os.path.join(d, 'DejaVuSans.ttf')
        bold = os.path.join(d, 'DejaVuSans-Bold.ttf')
        if os.path.exists(regular) and os.path.exists(bold):
            try:
                if 'DejaVuSans' not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont('DejaVuSans', regular))
                if 'DejaVuSans-Bold' not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', bold))
                # Map family so <b>/<strong> tags in Paragraph pick the bold variant
                registerFontFamily(
                    'DejaVuSans',
                    normal='DejaVuSans',
                    bold='DejaVuSans-Bold',
                    italic='DejaVuSans',
                    boldItalic='DejaVuSans-Bold',
                )
                _FONT_BASE = 'DejaVuSans'
                _FONT_BOLD = 'DejaVuSans-Bold'
                logger.info(f"PDF fonts registered from {d}")
                return
            except Exception as e:
                logger.warning(f"Failed to register fonts from {d}: {e}")
    logger.warning("DejaVuSans fonts not found — falling back to Helvetica (Czech diacritics may break)")

_register_pdf_fonts()


# ============ Czech IBAN generation ============

def cz_account_to_iban(account_number: str, bank_code: str) -> str:
    """Convert Czech BBAN to IBAN.
    account_number may be 'prefix-number' or just 'number'.
    Used for SPAYD QR payments so bank apps correctly parse the account.
    """
    if not account_number or not bank_code:
        return ''
    acc = str(account_number).strip()
    bank = str(bank_code).strip().rjust(4, '0')[-4:]
    if '-' in acc:
        prefix, main = acc.split('-', 1)
    else:
        prefix, main = '0', acc
    prefix = prefix.rjust(6, '0')[-6:]
    main = main.rjust(10, '0')[-10:]
    bban = f"{bank}{prefix}{main}"
    # Check digits: replace CZ=12 35 and 00 placeholder, mod-97
    numeric = bban + '1235' + '00'
    checksum = 98 - (int(numeric) % 97)
    return f"CZ{checksum:02d}{bban}"

STATUS_LABELS = {
    'pending': 'Čeká na schválení',
    'approved': 'Schváleno',
    'rejected': 'Zamítnuto',
}

PAYMENT_LABELS = {
    'unpaid': 'Nezaplaceno',
    'pending': 'Čeká na platbu',
    'paid': 'Zaplaceno',
    'not_required': 'Platba není vyžadována',
}


def _build_field_label_map(form_fields: list) -> dict:
    """Map field IDs to labels."""
    return {f.get('id', ''): f.get('label', f.get('id', '')) for f in (form_fields or [])}


def _format_date(iso: str) -> str:
    if not iso:
        return ''
    try:
        dt = datetime.fromisoformat(iso.replace('Z', '+00:00'))
        return dt.strftime('%d.%m.%Y %H:%M')
    except Exception:
        return iso[:10] if len(iso) >= 10 else iso


def _get_app_rows(applications: list, field_map: dict) -> tuple:
    """Build header + data rows from applications."""
    # Fixed columns
    fixed_headers = ['Jméno', 'Email', 'Status', 'Platba', 'Částka', 'VS', 'Datum přihlášení']
    
    # Dynamic field columns (from form_fields)
    dynamic_keys = list(field_map.keys())
    dynamic_headers = [field_map[k] for k in dynamic_keys]
    
    headers = fixed_headers + dynamic_headers
    rows = []
    
    for app in applications:
        data = app.get('applicant_data', {}) or {}
        row = [
            app.get('applicant_name', ''),
            app.get('applicant_email', ''),
            STATUS_LABELS.get(app.get('status', ''), app.get('status', '')),
            PAYMENT_LABELS.get(app.get('payment_status', ''), app.get('payment_status', '')),
            f"{app.get('total_amount', 0)} Kč",
            app.get('variable_symbol', ''),
            _format_date(app.get('created_at', '')),
        ]
        for key in dynamic_keys:
            val = data.get(key, '')
            if isinstance(val, bool):
                val = 'Ano' if val else 'Ne'
            row.append(str(val))
        rows.append(row)
    
    return headers, rows


# ============ XLSX Export ============

def generate_xlsx(event: dict, applications: list) -> io.BytesIO:
    """Generate styled XLSX export of applications."""
    form_fields = event.get('form_fields', [])
    field_map = _build_field_label_map(form_fields)
    headers, rows = _get_app_rows(applications, field_map)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Přihlášky"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )
    
    paid_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    unpaid_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    pending_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    title_cell = ws.cell(row=1, column=1, value=f"Přihlášky — {event.get('name', 'Událost')}")
    title_cell.font = Font(bold=True, size=14, color="1E293B")
    title_cell.alignment = Alignment(horizontal="left")
    
    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
    sub_cell = ws.cell(row=2, column=1, value=f"Exportováno: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Celkem: {len(applications)} přihlášek")
    sub_cell.font = Font(size=10, color="64748B")
    
    # Headers (row 4)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Data rows
    for row_idx, row_data in enumerate(rows, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Conditional formatting for payment status column (index 3, col 4)
            if col_idx == 4:
                if value == PAYMENT_LABELS['paid']:
                    cell.fill = paid_fill
                    cell.font = Font(color="166534", bold=True)
                elif value == PAYMENT_LABELS['unpaid']:
                    cell.fill = unpaid_fill
                    cell.font = Font(color="991B1B")
                elif value == PAYMENT_LABELS['pending']:
                    cell.fill = pending_fill
                    cell.font = Font(color="92400E")
        
        # Alternate row shading
        if row_idx % 2 == 0:
            for col_idx in range(1, len(row_data) + 1):
                if ws.cell(row=row_idx, column=col_idx).fill == PatternFill():
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                        start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"
                    )
    
    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row_idx in range(5, 5 + len(rows)):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)
    
    # Summary row
    summary_row = 5 + len(rows) + 1
    ws.cell(row=summary_row, column=1, value="Celkem").font = Font(bold=True)
    
    paid_count = sum(1 for a in applications if a.get('payment_status') == 'paid')
    total_amount = sum(a.get('total_amount', 0) for a in applications)
    ws.cell(row=summary_row, column=4, value=f"Zaplaceno: {paid_count}/{len(applications)}").font = Font(bold=True)
    ws.cell(row=summary_row, column=5, value=f"{total_amount} Kč").font = Font(bold=True)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============ CSV Export ============

def generate_csv(event: dict, applications: list) -> io.BytesIO:
    """Generate CSV export with UTF-8 BOM for Excel compatibility."""
    form_fields = event.get('form_fields', [])
    field_map = _build_field_label_map(form_fields)
    headers, rows = _get_app_rows(applications, field_map)
    
    buffer = io.BytesIO()
    # UTF-8 BOM for Excel
    buffer.write(b'\xef\xbb\xbf')
    
    # Write CSV content as bytes directly
    lines = []
    def csv_escape(val):
        s = str(val).replace('"', '""')
        if ';' in s or '"' in s or '\n' in s:
            return f'"{s}"'
        return s
    
    lines.append(';'.join(csv_escape(h) for h in headers))
    for row in rows:
        lines.append(';'.join(csv_escape(v) for v in row))
    
    buffer.write('\n'.join(lines).encode('utf-8'))
    buffer.seek(0)
    return buffer


# ============ PDF Confirmation ============

def generate_pdf_confirmation(
    application: dict,
    event: dict,
    event_date: Optional[dict],
    institution: dict,
    payment_settings: Optional[dict],
) -> io.BytesIO:
    """Generate PDF confirmation for a single application with QR payment."""
    buffer = io.BytesIO()
    
    base_font = _FONT_BASE
    bold_font = _FONT_BOLD
    
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )
    
    styles = getSampleStyleSheet()
    
    styles.add(ParagraphStyle(
        name='InstitutionName',
        fontSize=18, leading=22, spaceAfter=4,
        textColor=colors.HexColor('#1E293B'),
        fontName=bold_font,
    ))
    styles.add(ParagraphStyle(
        name='DocTitle',
        fontSize=14, leading=18, spaceAfter=12, spaceBefore=8,
        textColor=colors.HexColor('#334155'),
        fontName=bold_font,
    ))
    styles.add(ParagraphStyle(
        name='SectionTitle',
        fontSize=11, leading=14, spaceAfter=6, spaceBefore=12,
        textColor=colors.HexColor('#1E293B'),
        fontName=bold_font,
    ))
    styles.add(ParagraphStyle(
        name='BodyText2',
        fontSize=10, leading=14, spaceAfter=4,
        textColor=colors.HexColor('#475569'),
        fontName=base_font,
    ))
    styles.add(ParagraphStyle(
        name='Footer',
        fontSize=8, leading=10,
        textColor=colors.HexColor('#94A3B8'),
        alignment=1,
        fontName=base_font,
    ))
    
    elements = []
    
    # ── Header ──
    inst_name = institution.get('name', 'Instituce')
    elements.append(Paragraph(inst_name, styles['InstitutionName']))
    elements.append(Spacer(1, 2*mm))
    
    # Divider line
    elements.append(Table(
        [['']],
        colWidths=[170*mm],
        style=TableStyle([
            ('LINEBELOW', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0')),
        ])
    ))
    elements.append(Spacer(1, 4*mm))
    
    # ── Title ──
    elements.append(Paragraph("Potvrzení přihlášky", styles['DocTitle']))
    
    # ── Event info ──
    elements.append(Paragraph("Událost", styles['SectionTitle']))
    
    event_data = [
        ['Název:', event.get('name', '')],
        ['Typ:', {'event': 'Akce', 'camp': 'Příměstský tábor', 'workshop': 'Workshop', 'course': 'Kurz'}.get(event.get('type', ''), event.get('type', ''))],
    ]
    if event_date:
        event_data.append(['Termín:', f"{_format_date(event_date.get('start_datetime', ''))} — {_format_date(event_date.get('end_datetime', ''))}"])
    if event.get('price', 0) > 0:
        event_data.append(['Cena:', f"{event.get('price', 0)} {event.get('currency', 'Kč')}"])
    
    event_table = Table(event_data, colWidths=[35*mm, 135*mm])
    event_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), base_font),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748B')),
        ('FONTNAME', (1, 0), (1, -1), bold_font),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1E293B')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(event_table)
    
    # ── Applicant info ──
    elements.append(Spacer(1, 4*mm))
    elements.append(Paragraph("Údaje účastníka", styles['SectionTitle']))
    
    applicant_rows = []
    if application.get('applicant_name'):
        applicant_rows.append(['Jméno:', application['applicant_name']])
    if application.get('applicant_email'):
        applicant_rows.append(['Email:', application['applicant_email']])
    
    # Dynamic form fields
    field_map = _build_field_label_map(event.get('form_fields', []))
    app_data = application.get('applicant_data', {}) or {}
    for key, val in app_data.items():
        label = field_map.get(key, key)
        if isinstance(val, bool):
            val = 'Ano' if val else 'Ne'
        applicant_rows.append([f"{label}:", str(val)])
    
    if applicant_rows:
        app_table = Table(applicant_rows, colWidths=[45*mm, 125*mm])
        app_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), base_font),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748B')),
            ('FONTNAME', (1, 0), (1, -1), base_font),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1E293B')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        elements.append(app_table)
    
    # ── Payment section ──
    total = application.get('total_amount', 0)
    vs = application.get('variable_symbol', '')
    
    if total and total > 0 and payment_settings:
        elements.append(Spacer(1, 6*mm))
        elements.append(Paragraph("Platební údaje", styles['SectionTitle']))
        
        acc_num = payment_settings.get('account_number', '')
        bank_code = payment_settings.get('bank_code', '')
        acc_name = payment_settings.get('account_name', '')
        
        pay_rows = []
        if acc_num:
            pay_rows.append(['Číslo účtu:', f"{acc_num}/{bank_code}" if bank_code else acc_num])
        if acc_name:
            pay_rows.append(['Příjemce:', acc_name])
        pay_rows.append(['Částka:', f"{total} Kč"])
        pay_rows.append(['Variabilní symbol:', vs])
        
        pay_table = Table(pay_rows, colWidths=[45*mm, 125*mm])
        pay_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), base_font),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748B')),
            ('FONTNAME', (1, 0), (1, -1), bold_font),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1E293B')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ]))
        elements.append(pay_table)
        
        # QR code
        if acc_num and vs:
            try:
                import qrcode
                from reportlab.lib.utils import ImageReader
                
                iban = cz_account_to_iban(acc_num, bank_code)
                if not iban:
                    raise ValueError("Invalid account/bank code for IBAN")
                spayd = f"SPD*1.0*ACC:{iban}*AM:{total:.2f}*CC:CZK*X-VS:{vs}*MSG:Prihlaska"
                
                qr = qrcode.QRCode(version=1, box_size=6, border=2)
                qr.add_data(spayd)
                qr.make(fit=True)
                qr_img = qr.make_image(fill_color="black", back_color="white")
                
                qr_buffer = io.BytesIO()
                qr_img.save(qr_buffer, format='PNG')
                qr_buffer.seek(0)
                
                elements.append(Spacer(1, 4*mm))
                elements.append(Paragraph("QR platba:", styles['BodyText2']))
                qr_rl = RLImage(qr_buffer, width=40*mm, height=40*mm)
                elements.append(qr_rl)
                elements.append(Paragraph("Naskenujte QR kód v bankovní aplikaci.", styles['BodyText2']))
            except Exception as e:
                logger.warning(f"QR generation failed: {e}")
    
    # ── Status ──
    elements.append(Spacer(1, 6*mm))
    status_text = STATUS_LABELS.get(application.get('status', ''), application.get('status', ''))
    payment_text = PAYMENT_LABELS.get(application.get('payment_status', ''), application.get('payment_status', ''))
    elements.append(Paragraph(f"Status přihlášky: <b>{status_text}</b>  |  Platba: <b>{payment_text}</b>", styles['BodyText2']))
    
    # ── Footer ──
    elements.append(Spacer(1, 10*mm))
    elements.append(Table(
        [['']],
        colWidths=[170*mm],
        style=TableStyle([('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0'))])
    ))
    elements.append(Spacer(1, 3*mm))
    elements.append(Paragraph(
        f"Vygenerováno systémem Budeživo.cz | {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        styles['Footer']
    ))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer



# ==================================================================================
# Archive report & GDPR export PDFs — called from routes/programs.py and routes/gdpr.py
# ==================================================================================

def _pdf_base_styles():
    """Shared style-sheet for structured reports."""
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='RptTitle', fontName=_FONT_BOLD, fontSize=18,
                              textColor=colors.HexColor('#1F2937'), spaceAfter=6))
    styles.add(ParagraphStyle(name='RptSub', fontName=_FONT_BASE, fontSize=10,
                              textColor=colors.HexColor('#6B7280'), spaceAfter=10))
    styles.add(ParagraphStyle(name='RptH2', fontName=_FONT_BOLD, fontSize=13,
                              textColor=colors.HexColor('#111827'),
                              spaceBefore=10, spaceAfter=6))
    styles.add(ParagraphStyle(name='RptBody', fontName=_FONT_BASE, fontSize=10,
                              textColor=colors.HexColor('#1F2937'), leading=14))
    styles.add(ParagraphStyle(name='RptFooter', fontName=_FONT_BASE, fontSize=8,
                              textColor=colors.HexColor('#9CA3AF'), alignment=1))
    return styles


def _kv_table(rows, col_widths=(55*mm, 115*mm)):
    """Render a key/value table with clean light styling."""
    t = Table([[k, str(v) if v is not None else '—'] for k, v in rows], colWidths=col_widths)
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), _FONT_BASE),
        ('FONTNAME', (0, 0), (0, -1), _FONT_BOLD),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#111827')),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F9FAFB')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LINEBELOW', (0, 0), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
    ]))
    return t


def _footer(styles, label="Vygenerováno systémem Budeživo.cz"):
    return Paragraph(f"{label} | {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['RptFooter'])


def _resolve_local_image(url_or_path: Optional[str]) -> Optional[str]:
    """Return a local filesystem path for a program cover image, or None.

    Supports:
      * absolute filesystem paths (`/app/uploads/...`)
      * URL-style paths starting with `/uploads/` or `/static/` — mapped to
        the corresponding local directory.

    Remote URLs (http/https/s3) are downloaded into a tempfile so reportlab
    can embed them. Failures return None — the caller must handle gracefully.
    """
    if not url_or_path:
        return None
    import os
    import tempfile
    p = str(url_or_path).strip()
    if not p:
        return None
    if p.startswith("/") and not p.startswith("//"):
        if os.path.exists(p):
            return p
        backend_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        candidate = os.path.join(backend_root, p.lstrip("/"))
        if os.path.exists(candidate):
            return candidate
        return None
    if p.startswith("http://") or p.startswith("https://"):
        try:
            import urllib.request
            suffix = ".jpg"
            for ext in (".png", ".jpg", ".jpeg", ".webp"):
                if p.lower().split("?")[0].endswith(ext):
                    suffix = ext
                    break
            req = urllib.request.Request(p, headers={"User-Agent": "BudeZivo-PDF/1.0"})
            with urllib.request.urlopen(req, timeout=8) as resp:  # noqa: S310 — internal use
                data = resp.read()
            tmp = tempfile.NamedTemporaryFile(prefix="bz_cover_", suffix=suffix, delete=False)
            tmp.write(data)
            tmp.close()
            return tmp.name
        except Exception as e:
            logger.warning(f"Could not fetch hero image {p!r}: {e}")
            return None
    return None


def _build_hero_cover(prog: dict, inst: dict, stats: dict, image_path: str, styles) -> list:
    """Build the cover page for the HERO archive PDF variant.

    Layout: a large image (about 70 % of page height) followed by an overlay-
    style dark band carrying program identifiers, then a forced page break.
    A reportlab ``SimpleDocTemplate`` cannot render true bleed-to-edge content
    inside a flowable frame, so we accept the standard top/bottom margin.
    """
    from reportlab.platypus import PageBreak
    from reportlab.lib.pagesizes import A4 as _A4

    page_w, page_h = _A4
    available_w = page_w - (40 * mm)  # 20 mm margins on both sides

    cover_styles = {
        "kicker": ParagraphStyle(
            "HeroKicker", parent=styles['RptBody'], textColor=colors.white,
            fontSize=10, leading=12, fontName=_FONT_BOLD, alignment=0,
        ),
        "title": ParagraphStyle(
            "HeroTitle", parent=styles['RptTitle'], textColor=colors.white,
            fontSize=30, leading=36, fontName=_FONT_BOLD, alignment=0,
        ),
        "meta": ParagraphStyle(
            "HeroMeta", parent=styles['RptBody'], textColor=colors.white,
            fontSize=11, leading=15, fontName=_FONT_BASE, alignment=0,
        ),
    }

    # Image sized to ~ 65 % of usable page height; reportlab preserves aspect
    # ratio with `kind='proportional'` only when both width and height fit.
    target_h = (page_h - (35 * mm)) * 0.65
    try:
        cover_img = RLImage(image_path, width=available_w, height=target_h, kind='proportional')
    except Exception as e:
        logger.warning(f"Hero image render failed: {e}")
        return []

    kicker = "DOPROVODNÝ PROGRAM — ARCHIVNÍ ZPRÁVA"
    title_html = (prog.get('name') or '—').replace('&', '&amp;')

    meta_lines = []
    if prog.get('age_group'):
        meta_lines.append(f"<b>{prog['age_group']}</b>")
    if inst.get('name'):
        meta_lines.append(inst['name'])
    dr = stats.get('date_range', {}) or {}
    if dr.get('from') or dr.get('to'):
        meta_lines.append(f"Období: {dr.get('from') or '—'} → {dr.get('to') or '—'}")

    band = Table(
        [
            [Paragraph(kicker, cover_styles["kicker"])],
            [Paragraph(title_html, cover_styles["title"])],
            [Paragraph("<br/>".join(meta_lines) or "&nbsp;", cover_styles["meta"])],
        ],
        colWidths=[available_w],
    )
    band.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#263FA8')),
        ('LEFTPADDING', (0, 0), (-1, -1), 14),
        ('RIGHTPADDING', (0, 0), (-1, -1), 14),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))

    return [cover_img, Spacer(1, 4 * mm), band, PageBreak()]


def build_archive_report_pdf(data: dict, custom_text: Optional[str] = None) -> bytes:
    """Render a program archive report as a premium editorial-style A4 PDF.

    The implementation has been migrated to ``services.pdf`` (a dedicated PDF
    render layer with cover hero, KPI dashboard, minimal charts, gallery and
    quote cards). The function signature, callers and endpoint behaviour are
    unchanged for backward compatibility.
    """
    from services.pdf import render_program_report
    return render_program_report(data, custom_text=custom_text)


def _legacy_build_archive_report_pdf(data: dict, custom_text: Optional[str] = None) -> bytes:  # noqa: C901
    """Legacy renderer kept only as an emergency fallback. Not exported."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=18*mm, bottomMargin=15*mm)
    styles = _pdf_base_styles()
    story = []

    prog = data.get('program', {}) or {}
    inst = data.get('institution', {}) or {}
    stats = data.get('statistics', {}) or {}

    # ---- HERO cover page (only when a cover image is available) ----
    cover_path = _resolve_local_image(prog.get('image_url'))
    if cover_path:
        story.extend(_build_hero_cover(prog, inst, stats, cover_path, styles))

    story.append(Paragraph(f"Archive report – {prog.get('name', '—')}", styles['RptTitle']))
    story.append(Paragraph(
        f"{inst.get('name', '—')} &nbsp;·&nbsp; vygenerováno {data.get('report_generated_at', '')[:19]}",
        styles['RptSub'],
    ))

    story.append(Paragraph("Přehled programu", styles['RptH2']))
    story.append(_kv_table([
        ("Název", prog.get('name')),
        ("Věková kategorie", prog.get('age_group')),
        ("Délka", f"{prog.get('duration')} min" if prog.get('duration') else None),
        ("Kapacita", prog.get('capacity')),
        ("Cenový info", prog.get('pricing_info')),
        ("Status", prog.get('status')),
        ("Archivováno", prog.get('archived_at')),
        ("Důvod archivace", prog.get('archive_reason')),
    ]))

    # ---- Description ("O programu — co se žáci naučí") ----
    if prog.get('description'):
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph("O programu — co se žáci naučí", styles['RptH2']))
        # Preserve user line-breaks; allow simple inline HTML.
        desc_html = (prog['description'] or '').replace('\n', '<br/>')
        story.append(Paragraph(desc_html, styles['RptBody']))

    # ---- Custom curatorial note (free-form, optional) ----
    if custom_text and custom_text.strip():
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph("Poznámka", styles['RptH2']))
        note_html = custom_text.strip().replace('&', '&amp;').replace('\n', '<br/>')
        story.append(Paragraph(note_html, styles['RptBody']))

    story.append(Paragraph("Statistiky", styles['RptH2']))
    dr = stats.get('date_range', {}) or {}
    story.append(_kv_table([
        ("Rezervací celkem", stats.get('total_reservations', 0)),
        ("Potvrzené / dokončené / zrušené",
         f"{stats.get('confirmed', 0)} / {stats.get('completed', 0)} / {stats.get('cancelled', 0)}"),
        ("Studentů celkem", stats.get('total_students', 0)),
        ("Pedagogů celkem", stats.get('total_teachers', 0)),
        ("Unikátních škol", stats.get('unique_schools', 0)),
        ("Rozsah dat", f"{dr.get('from') or '—'} → {dr.get('to') or '—'}"),
        ("Zpětných vazeb", data.get('feedback_count', 0)),
    ]))

    schools = data.get('schools', {}) or {}
    if schools:
        story.append(Paragraph("Školy", styles['RptH2']))
        rows = [["Škola", "Návštěv", "Studentů", "Poslední"]]
        for sn, info in sorted(schools.items(), key=lambda x: -x[1].get('visits', 0))[:25]:
            rows.append([sn[:60], info.get('visits', 0), info.get('students', 0),
                         info.get('last_visit') or '—'])
        t = Table(rows, colWidths=(90*mm, 20*mm, 25*mm, 35*mm))
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), _FONT_BASE),
            ('FONTNAME', (0, 0), (-1, 0), _FONT_BOLD),
            ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F3F4F6')),
            ('LINEBELOW', (0, 0), (-1, 0), 0.75, colors.HexColor('#6B7280')),
            ('LINEBELOW', (0, 1), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
            ('ALIGN', (1, 1), (3, -1), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t)

    bookings = data.get('bookings', []) or []
    if bookings:
        story.append(Paragraph(f"Rezervace ({len(bookings)})", styles['RptH2']))
        rows = [["Datum", "Čas", "Škola", "Status", "Studentů"]]
        for b in bookings[:80]:
            rows.append([b.get('date') or '—', b.get('time_block') or '—',
                         (b.get('school_name') or '—')[:42],
                         b.get('status') or '—', b.get('num_students') or '—'])
        t = Table(rows, colWidths=(22*mm, 25*mm, 65*mm, 30*mm, 18*mm))
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), _FONT_BASE),
            ('FONTNAME', (0, 0), (-1, 0), _FONT_BOLD),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F3F4F6')),
            ('LINEBELOW', (0, 0), (-1, 0), 0.75, colors.HexColor('#6B7280')),
            ('LINEBELOW', (0, 1), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t)
        if len(bookings) > 80:
            story.append(Paragraph(f"<i>… zkráceno, zobrazeno prvních 80 z {len(bookings)} rezervací.</i>",
                                   styles['RptBody']))

    story.append(Spacer(1, 6*mm))
    story.append(_footer(styles))
    doc.build(story)
    return buf.getvalue()


def build_gdpr_export_pdf(data: dict) -> bytes:
    """Render the GDPR portability export (JSON dict) as a human-readable PDF.
    Note: this is a *companion* document; the authoritative machine-readable format
    for GDPR Article 20 remains JSON, which is delivered alongside."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=18*mm, bottomMargin=15*mm)
    styles = _pdf_base_styles()
    story = []

    story.append(Paragraph("Export osobních údajů (GDPR čl. 20)", styles['RptTitle']))
    story.append(Paragraph(
        f"Datum exportu: {data.get('export_date', '')[:19]} &nbsp;·&nbsp; "
        "Pro přenositelnost v souladu s GDPR použijte JSON variantu tohoto souboru.",
        styles['RptSub'],
    ))

    u = data.get('user_data', {}) or {}
    story.append(Paragraph("Uživatel", styles['RptH2']))
    story.append(_kv_table([
        ("ID", u.get('id')),
        ("E-mail", u.get('email')),
        ("Jméno", u.get('name')),
        ("Role", u.get('role')),
        ("Status", u.get('status')),
        ("GDPR souhlas", u.get('gdpr_consent')),
        ("Datum souhlasu", u.get('gdpr_consent_date')),
        ("Účet vytvořen", u.get('created_at')),
    ]))

    i = data.get('institution_data', {}) or {}
    if i:
        story.append(Paragraph("Instituce", styles['RptH2']))
        story.append(_kv_table([
            ("Název", i.get('name')),
            ("Typ", i.get('type')),
            ("Adresa", f"{i.get('address') or ''}, {i.get('city') or ''}, {i.get('country') or ''}".strip(' ,')),
            ("Plán", i.get('plan')),
            ("Vytvořeno", i.get('created_at')),
        ]))

    bookings = data.get('bookings', []) or []
    story.append(Paragraph(f"Rezervace ({data.get('bookings_count', len(bookings))})", styles['RptH2']))
    if bookings:
        rows = [["Datum", "Škola", "Kontakt", "Status", "Žáci"]]
        for b in bookings[:80]:
            rows.append([
                b.get('date') or '—',
                (b.get('school_name') or '—')[:35],
                (f"{b.get('contact_name') or ''}\n{b.get('contact_email') or ''}").strip(),
                b.get('status') or '—',
                b.get('num_students') or '—',
            ])
        t = Table(rows, colWidths=(22*mm, 50*mm, 62*mm, 25*mm, 15*mm))
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), _FONT_BASE),
            ('FONTNAME', (0, 0), (-1, 0), _FONT_BOLD),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F3F4F6')),
            ('LINEBELOW', (0, 0), (-1, 0), 0.75, colors.HexColor('#6B7280')),
            ('LINEBELOW', (0, 1), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t)
        if len(bookings) > 80:
            story.append(Paragraph(f"<i>… zkráceno, zobrazeno prvních 80 z {len(bookings)}.</i>",
                                   styles['RptBody']))
    else:
        story.append(Paragraph("Žádné rezervace.", styles['RptBody']))

    schools = data.get('schools', []) or []
    story.append(Paragraph(f"Školy v databázi ({data.get('schools_count', len(schools))})", styles['RptH2']))
    if schools:
        rows = [["Název", "Kontakt", "E-mail", "Město"]]
        for s in schools[:80]:
            rows.append([(s.get('name') or '—')[:40], s.get('contact_person') or '—',
                         s.get('email') or '—', s.get('city') or '—'])
        t = Table(rows, colWidths=(60*mm, 45*mm, 45*mm, 25*mm))
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), _FONT_BASE),
            ('FONTNAME', (0, 0), (-1, 0), _FONT_BOLD),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F3F4F6')),
            ('LINEBELOW', (0, 0), (-1, 0), 0.75, colors.HexColor('#6B7280')),
            ('LINEBELOW', (0, 1), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("Žádné školy.", styles['RptBody']))

    story.append(Spacer(1, 6*mm))
    story.append(_footer(styles, label="GDPR export — Budeživo.cz"))
    doc.build(story)
    return buf.getvalue()
